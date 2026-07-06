from __future__ import annotations

import html
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
REFERENCE_XLSX_PATH = ROOT / "KN機械査定_参考.xlsx"
PHOTO_DIR = ROOT / "写真"
OUTPUT_PATH = ROOT / "index.html"
DEFAULT_TITLE = "機械査定"
DEFAULT_SUBTITLE = "（有）河津内装　殿"
DEFAULT_NOTE = "PDFの印刷部分のみ転記（手書きは未反映）"
DEFAULT_SUMMARY_DATE = "2026年6月29日"
EXCLUDED_NUMBERS = set()
SOLD_REMARK = "売約済み"
SOLD_NUMBERS = {22, 23, 33}


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def find_photo(hyperlink: str, row_no: int) -> str:
    photo_index = {path.name.lower(): path.name for path in PHOTO_DIR.glob("*") if path.is_file()}

    candidate_names: list[str] = []
    if hyperlink:
        candidate_names.append(Path(hyperlink).name)
    candidate_names.extend([f"{row_no}.JPG", f"{row_no}.jpg", f"{row_no}.jpeg", f"{row_no}.JPEG"])

    for candidate in candidate_names:
        actual_name = photo_index.get(candidate.lower())
        if actual_name:
            return f"写真/{actual_name}"
    return ""


def build_reference_rows() -> dict[int, dict[str, str]]:
    if not REFERENCE_XLSX_PATH.exists():
        raise FileNotFoundError(f"Source workbook not found: {REFERENCE_XLSX_PATH.name}")

    workbook = load_workbook(REFERENCE_XLSX_PATH, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    rows: dict[int, dict[str, str]] = {}
    for row in range(1, worksheet.max_row + 1):
        number = worksheet[f"A{row}"].value
        if not isinstance(number, int):
            continue
        rows[number] = {
            "machineName": clean(worksheet[f"B{row}"].value),
            "quantity": clean(worksheet[f"C{row}"].value),
            "remarks": clean(worksheet[f"F{row}"].value),
        }
    return rows


def build_data() -> dict[str, object]:
    reference_rows = build_reference_rows()

    items: list[dict[str, object]] = []
    for number in sorted(reference_rows):
        if number in EXCLUDED_NUMBERS:
            continue

        photo_path = find_photo("", number)
        reference_row = reference_rows[number]
        reference_name = reference_row.get("machineName", "")
        reference_quantity = reference_row.get("quantity", "")
        reference_remarks = reference_row.get("remarks", "")

        items.append(
            {
                "no": number,
                "machineName": reference_name,
                "quantity": reference_quantity,
                "desiredPrice": "",
                "remarks": SOLD_REMARK if number in SOLD_NUMBERS or reference_remarks else "",
                "photoPath": photo_path,
            }
        )

    return {
        "title": DEFAULT_TITLE,
        "subtitle": DEFAULT_SUBTITLE,
        "note": DEFAULT_NOTE,
        "summaryDate": DEFAULT_SUMMARY_DATE,
        "items": items,
    }


def render_html(data: dict[str, object]) -> str:
    payload = json.dumps(data, ensure_ascii=False)
    page_title = html.escape(str(data["title"]))
    template = """<!DOCTYPE html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>__PAGE_TITLE__</title>
  <style>
    :root {{
      --bg: #f3efe7;
      --panel: rgba(255, 255, 255, 0.88);
      --panel-strong: #fffdf8;
      --ink: #1f2a27;
      --muted: #5d6b67;
      --accent: #0f766e;
      --accent-strong: #0a4f4a;
      --line: rgba(20, 32, 29, 0.12);
      --shadow: 0 24px 70px rgba(31, 42, 39, 0.14);
    }}

    * {{ box-sizing: border-box; }}

    body {{
      margin: 0;
      font-family: "Yu Gothic", "Hiragino Kaku Gothic ProN", sans-serif;
      color: var(--ink);
      background:
        radial-gradient(circle at top left, rgba(15, 118, 110, 0.18), transparent 32%),
        linear-gradient(180deg, #f4f0e8 0%, #ebe4d7 100%);
      min-height: 100vh;
    }}

    .shell {{
      width: min(1240px, calc(100% - 32px));
      margin: 0 auto;
      padding: 32px 0 48px;
    }}

    .hero {{
      background: linear-gradient(135deg, rgba(255,255,255,0.92), rgba(247, 242, 232, 0.84));
      border: 1px solid rgba(255,255,255,0.82);
      border-radius: 28px;
      box-shadow: var(--shadow);
      padding: 28px;
      backdrop-filter: blur(10px);
    }}

    .eyebrow {{
      display: inline-block;
      padding: 6px 12px;
      border-radius: 999px;
      background: rgba(15, 118, 110, 0.1);
      color: var(--accent-strong);
      font-size: 12px;
      font-weight: 700;
      letter-spacing: 0.08em;
    }}

    h1 {{
      margin: 14px 0 10px;
      font-size: clamp(28px, 4vw, 44px);
      line-height: 1.1;
    }}

    .lede,
    .meta {{
      margin: 0;
      color: var(--muted);
      line-height: 1.7;
    }}

    .hero-grid {{
      display: grid;
      grid-template-columns: 1.4fr 0.8fr;
      gap: 18px;
      margin-top: 20px;
    }}

    .hero-card {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 22px;
      padding: 18px;
    }}

    .hero-card strong {{
      display: block;
      margin-bottom: 8px;
      font-size: 14px;
    }}

    .toolbar {{
      display: grid;
      grid-template-columns: 1.4fr 0.8fr auto;
      gap: 14px;
      align-items: center;
      margin: 22px 0 18px;
    }}

    .search,
    .filter {{
      width: 100%;
      padding: 14px 16px;
      border-radius: 16px;
      border: 1px solid var(--line);
      background: rgba(255,255,255,0.82);
      font-size: 15px;
      color: var(--ink);
    }}

    .count {{
      text-align: right;
      color: var(--muted);
      font-size: 14px;
      font-weight: 700;
    }}

    .gallery {{
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(240px, 1fr));
      gap: 18px;
    }}

    .card {{
      background: var(--panel-strong);
      border: 1px solid var(--line);
      border-radius: 24px;
      overflow: hidden;
      box-shadow: var(--shadow);
    }}

    .card-button {{
      display: block;
      width: 100%;
      border: none;
      padding: 0;
      background: none;
      cursor: pointer;
      text-align: left;
    }}

    .card-photo {{
      aspect-ratio: 4 / 3;
      background: #dde4df;
      display: flex;
      align-items: center;
      justify-content: center;
      overflow: hidden;
    }}

    .card-photo img {{
      width: 100%;
      height: 100%;
      object-fit: cover;
      display: block;
    }}

    .card-photo.is-empty {{
      padding: 20px;
      color: var(--muted);
      font-size: 14px;
      line-height: 1.6;
    }}

    .card-body {{
      padding: 16px 16px 18px;
    }}

    .card-no {{
      margin: 0 0 8px;
      color: var(--accent-strong);
      font-size: 12px;
      font-weight: 700;
      letter-spacing: 0.06em;
      text-transform: uppercase;
    }}

    .card-title {{
      margin: 0;
      font-size: 16px;
      line-height: 1.55;
      color: var(--ink);
    }}

    .card.is-sold .card-title {{
      color: #8b928f;
    }}

    .card-meta {{
      display: grid;
      gap: 10px;
      margin-top: 14px;
    }}

    .card-meta-row {{
      padding-top: 10px;
      border-top: 1px solid var(--line);
    }}

    .card-meta-row strong {{
      display: block;
      margin-bottom: 4px;
      color: var(--muted);
      font-size: 11px;
      letter-spacing: 0.05em;
      text-transform: uppercase;
    }}

    .card-meta-row span {{
      display: block;
      font-size: 14px;
      line-height: 1.55;
    }}

    .empty {{
      padding: 28px;
      text-align: center;
      color: var(--muted);
      background: rgba(255,255,255,0.74);
      border: 1px solid var(--line);
      border-radius: 24px;
      box-shadow: var(--shadow);
    }}

    .muted {{
      color: var(--muted);
      font-size: 13px;
    }}

    @media (max-width: 900px) {{
      .hero-grid,
      .toolbar {{
        grid-template-columns: 1fr;
      }}

      .count {{
        text-align: left;
      }}
    }}

    @media (max-width: 640px) {{
      .shell {{
        width: min(100% - 20px, 1240px);
        padding-top: 20px;
      }}

      .hero {{
        padding: 22px 18px;
        border-radius: 22px;
      }}

      .gallery {{
        grid-template-columns: 1fr;
      }}
    }}
  </style>
</head>
<body>
  <div class="shell">
    <section class="hero">
      <span class="eyebrow">MACHINE APPRAISAL LIST</span>
      <h1 id="page-title"></h1>
      <p class="lede" id="page-subtitle"></p>
      <div class="hero-grid">
        <div class="hero-card">
          <strong>ご利用方法</strong>
          <p class="meta">写真を一覧で見ながら機械を確認できます。写真をクリックすると大きく表示されます。</p>
        </div>
        <div class="hero-card">
          <strong>掲載情報</strong>
          <p class="meta" id="page-meta"></p>
        </div>
      </div>
    </section>

    <section class="toolbar" aria-label="検索と絞り込み">
      <input id="searchInput" class="search" type="search" placeholder="機械名・備考で検索">
      <select id="photoFilter" class="filter">
        <option value="all">すべて表示</option>
        <option value="with-photo">写真ありのみ</option>
        <option value="without-photo">写真なしのみ</option>
      </select>
      <div class="count" id="resultCount"></div>
    </section>

    <section id="gallery" class="gallery"></section>
    <div class="empty" id="emptyState" hidden>該当する機械がありません。</div>
  </div>

  <script>
    const siteData = __PAYLOAD__;
    const items = siteData.items;
    const els = {{
      title: document.getElementById("page-title"),
      subtitle: document.getElementById("page-subtitle"),
      meta: document.getElementById("page-meta"),
      search: document.getElementById("searchInput"),
      filter: document.getElementById("photoFilter"),
      count: document.getElementById("resultCount"),
      gallery: document.getElementById("gallery"),
      empty: document.getElementById("emptyState"),
    }};

    function escapeHtml(value) {{
      return String(value ?? "")
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#39;");
    }}

    function summarize() {{
      const countText = `掲載 ${items.length} 台`;
      const dateText = siteData.summaryDate ? ` / 作成日 ${siteData.summaryDate}` : "";
      const noteText = siteData.note ? ` / ${siteData.note}` : "";
      return countText + dateText + noteText;
    }}

    function openPhoto(item) {{
      if (!item.photoPath) {{
        window.alert("この機械には写真が登録されていません。");
        return;
      }}
      window.open(encodeURI(item.photoPath), "_blank", "noopener,noreferrer");
    }}

    function renderCards() {{
      const keyword = els.search.value.trim().toLowerCase();
      const photoMode = els.filter.value;
      const filtered = items.filter((item) => {{
        const text = [item.machineName, item.remarks].join(" ").toLowerCase();
        const matchKeyword = !keyword || text.includes(keyword);
        const matchPhoto =
          photoMode === "all" ||
          (photoMode === "with-photo" && item.photoPath) ||
          (photoMode === "without-photo" && !item.photoPath);
        return matchKeyword && matchPhoto;
      }});

      els.count.textContent = `${filtered.length} 件表示`;
      els.empty.hidden = filtered.length !== 0;

      els.gallery.innerHTML = filtered.map((item) => {{
        const photoBlock = item.photoPath
          ? `<div class="card-photo"><img src="${{encodeURI(item.photoPath)}}" alt="${{escapeHtml(item.machineName)}}"></div>`
          : `<div class="card-photo is-empty">写真未登録</div>`;
        const soldClass = String(item.remarks || "").includes("売約済み") ? " is-sold" : "";

        return `
          <article class="card${{soldClass}}">
            <button class="card-button" type="button" data-no="${{escapeHtml(item.no)}}" aria-label="${{escapeHtml(item.machineName)}} の写真を開く">
              ${photoBlock}
              <div class="card-body">
                <p class="card-no">No. ${{escapeHtml(item.no)}}</p>
                <h2 class="card-title">${{escapeHtml(item.machineName)}}</h2>
                <div class="card-meta">
                  <div class="card-meta-row"><strong>数量</strong><span>${{escapeHtml(item.quantity || "-")}}</span></div>
                  <div class="card-meta-row"><strong>備考</strong><span>${{escapeHtml(item.remarks || "-")}}</span></div>
                </div>
              </div>
            </button>
          </article>
        `;
      }}).join("");

      for (const button of els.gallery.querySelectorAll(".card-button")) {{
        button.addEventListener("click", () => {{
          const item = items.find((entry) => String(entry.no) === button.dataset.no);
          if (item) openPhoto(item);
        }});
      }}
    }}

    els.title.textContent = siteData.title;
    els.subtitle.textContent = siteData.subtitle || "写真の下に機械名が並ぶ一覧です。";
    els.meta.textContent = summarize();

    els.search.addEventListener("input", renderCards);
    els.filter.addEventListener("change", renderCards);
    renderCards();
  </script>
</body>
</html>
"""
    return template.replace("__PAGE_TITLE__", page_title).replace("__PAYLOAD__", payload).replace("{{", "{").replace("}}", "}")


def main() -> None:
    data = build_data()
    OUTPUT_PATH.write_text(render_html(data), encoding="utf-8")
    print(f"Generated: {OUTPUT_PATH}")
    print(f"Items: {len(data['items'])}")


if __name__ == "__main__":
    main()
